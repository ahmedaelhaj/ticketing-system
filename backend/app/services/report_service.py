import io
from datetime import datetime
from xml.sax.saxutils import escape as _esc

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo

from app.models import Ticket

BRAND = "#4338CA"
INK = "#1E293B"
MUTED = "#64748B"
GRID = "#E2E8F0"
STRIPE = "#F8FAFC"

STATUS_COLORS = {
    "pending_approval": ("#F59E0B", "#FEF3C7"),
    "open": ("#2563EB", "#DBEAFE"),
    "in_progress": ("#4F46E5", "#E0E7FF"),
    "closed": ("#059669", "#D1FAE5"),
    "rejected": ("#E11D48", "#FFE4E6"),
}
STATUS_LABELS = {
    "pending_approval": "Pending approval", "open": "Open", "in_progress": "In progress",
    "closed": "Closed", "rejected": "Rejected",
}
PRIORITY_ORDER = ["urgent", "high", "medium", "low"]

HEADER = ["Title", "Status", "Priority", "Requested by", "Origin team", "Current team",
          "Assignee", "Closed by", "Created", "Closed"]


def _row(t: Ticket) -> list:
    return [
        t.title,
        t.status.value,
        t.priority.value.title(),
        t.creator.full_name if t.creator else "-",
        t.origin_team.name if getattr(t, "origin_team", None) else (t.team.name if t.team else "-"),
        t.team.name if t.team else "-",
        t.assignee.full_name if t.assignee else "-",
        t.closer.full_name if getattr(t, "closer", None) else "-",
        t.created_at.strftime("%Y-%m-%d %H:%M"),
        t.closed_at.strftime("%Y-%m-%d %H:%M") if t.closed_at else "-",
    ]


def _summarize(tickets: list[Ticket]) -> tuple[dict, dict]:
    by_status: dict[str, int] = {}
    by_priority: dict[str, int] = {}
    for t in tickets:
        by_status[t.status.value] = by_status.get(t.status.value, 0) + 1
        by_priority[t.priority.value] = by_priority.get(t.priority.value, 0) + 1
    return by_status, by_priority


# ==================================================================== PDF ===

def _footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.HexColor(MUTED))
    canvas.drawString(1.5 * cm, 1 * cm, "SME Ticketing System")
    canvas.drawRightString(landscape(A4)[0] - 1.5 * cm, 1 * cm, f"Page {canvas.getPageNumber()}")
    canvas.setStrokeColor(colors.HexColor(GRID))
    canvas.line(1.5 * cm, 1.3 * cm, landscape(A4)[0] - 1.5 * cm, 1.3 * cm)
    canvas.restoreState()


def generate_pdf(tickets: list[Ticket], title: str, subtitle: str = "") -> bytes:
    buffer = io.BytesIO()
    page_width = landscape(A4)[0]
    usable_width = page_width - 3 * cm
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.3 * cm, bottomMargin=1.8 * cm,
                             leftMargin=1.5 * cm, rightMargin=1.5 * cm)

    styles = getSampleStyleSheet()
    eyebrow = ParagraphStyle("eyebrow", parent=styles["Normal"], textColor=colors.HexColor(BRAND),
                              fontSize=8.5, fontName="Helvetica-Bold", spaceAfter=2, leading=10)
    heading = ParagraphStyle("heading", parent=styles["Title"], textColor=colors.HexColor(INK),
                              fontSize=19, leading=22, spaceAfter=2)
    meta = ParagraphStyle("meta", parent=styles["Normal"], textColor=colors.HexColor(MUTED), fontSize=8.5)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.6, leading=9.2, textColor=colors.HexColor(INK))
    cell_muted = ParagraphStyle("cell_muted", parent=cell, textColor=colors.HexColor(MUTED))
    box_num = ParagraphStyle("box_num", parent=styles["Normal"], fontSize=15, fontName="Helvetica-Bold",
                              alignment=TA_CENTER, leading=17)
    box_label = ParagraphStyle("box_label", parent=styles["Normal"], fontSize=7, alignment=TA_CENTER,
                                leading=9)
    empty_style = ParagraphStyle("empty", parent=styles["Normal"], fontSize=10.5,
                                  textColor=colors.HexColor(MUTED), alignment=TA_CENTER)

    generated = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} · {len(tickets)} ticket(s)"
    meta_line = f"{generated} · {subtitle}" if subtitle else generated

    elements = [
        Paragraph("SME TICKETING SYSTEM", eyebrow),
        Paragraph(title, heading),
        Paragraph(meta_line, meta),
        Spacer(1, 0.45 * cm),
    ]

    if not tickets:
        elements.append(Spacer(1, 1.5 * cm))
        elements.append(Paragraph("No tickets match these filters.", empty_style))
        doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
        return buffer.getvalue()

    by_status, by_priority = _summarize(tickets)

    # --- summary stat boxes (one per status present) ---
    present_statuses = [s for s in STATUS_LABELS if by_status.get(s)]
    if present_statuses:
        box_w = usable_width / len(present_statuses)
        box_row = [[Paragraph(str(by_status[s]), box_num)] for s in present_statuses]
        label_row = [[Paragraph(STATUS_LABELS[s], box_label)] for s in present_statuses]
        stat_table = Table(
            [[c[0] for c in box_row], [c[0] for c in label_row]],
            colWidths=[box_w] * len(present_statuses),
            rowHeights=[22, 14],
        )
        style_cmds = [
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
        ]
        for i, s in enumerate(present_statuses):
            text_hex, bg_hex = STATUS_COLORS[s]
            style_cmds += [
                ("BACKGROUND", (i, 0), (i, 1), colors.HexColor(bg_hex)),
                ("TEXTCOLOR", (i, 0), (i, 0), colors.HexColor(text_hex)),
                ("TEXTCOLOR", (i, 1), (i, 1), colors.HexColor(text_hex)),
                ("LINEBELOW", (i, 1), (i, 1), 2, colors.HexColor(text_hex)),
            ]
        stat_table.setStyle(TableStyle(style_cmds))
        elements.append(stat_table)
        elements.append(Spacer(1, 0.1 * cm))

    priority_line = " · ".join(
        f"{p.title()}: {by_priority[p]}" for p in PRIORITY_ORDER if by_priority.get(p)
    )
    if priority_line:
        elements.append(Paragraph(f"By priority — {priority_line}", meta))
    elements.append(Spacer(1, 0.4 * cm))

    # --- main table ---
    col_weights = [2.6, 1.15, 0.85, 1.3, 1.1, 1.1, 1.3, 1.3, 1.15, 1.15]
    total_weight = sum(col_weights)
    col_widths = [usable_width * (w / total_weight) for w in col_weights]

    header_row = [Paragraph(h, ParagraphStyle("h", parent=cell, textColor=colors.white,
                                               fontName="Helvetica-Bold", fontSize=7.8)) for h in HEADER]
    body_rows = []
    for t in tickets:
        raw = _row(t)
        status_key = raw[1]
        text_hex, _ = STATUS_COLORS.get(status_key, ("#334155", "#F1F5F9"))
        row = [
            Paragraph(_esc(raw[0]), cell),
            Paragraph(STATUS_LABELS.get(status_key, status_key),
                      ParagraphStyle("s", parent=cell, textColor=colors.HexColor(text_hex), fontName="Helvetica-Bold")),
            Paragraph(_esc(raw[2]), cell),
            Paragraph(_esc(raw[3]), cell_muted),
            Paragraph(_esc(raw[4]), cell_muted),
            Paragraph(_esc(raw[5]), cell_muted),
            Paragraph(_esc(raw[6]), cell_muted),
            Paragraph(_esc(raw[7]), cell_muted),
            Paragraph(_esc(raw[8]), cell_muted),
            Paragraph(_esc(raw[9]), cell_muted),
        ]
        body_rows.append(row)

    rows = [header_row] + body_rows
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(INK)),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(GRID)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(STRIPE)]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


# ================================================================== Excel ===

def _style_header_row(ws, row_idx: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=INK.lstrip("#"))
        cell.alignment = Alignment(vertical="center")


def generate_excel(tickets: list[Ticket], title: str, subtitle: str = "") -> bytes:
    wb = Workbook()

    # ---------- Summary sheet ----------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = title
    ws_sum["A1"].font = Font(bold=True, size=16, color=BRAND.lstrip("#"))
    ws_sum["A2"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} · {len(tickets)} ticket(s)"
    ws_sum["A2"].font = Font(italic=True, size=9, color=MUTED.lstrip("#"))
    if subtitle:
        ws_sum["A3"] = subtitle
        ws_sum["A3"].font = Font(size=9, color=MUTED.lstrip("#"))

    by_status, by_priority = _summarize(tickets)
    row = 5
    ws_sum.cell(row=row, column=1, value="By status").font = Font(bold=True, size=11)
    row += 1
    ws_sum.cell(row=row, column=1, value="Status").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1
    for s in STATUS_LABELS:
        count = by_status.get(s, 0)
        if not count:
            continue
        text_hex, bg_hex = STATUS_COLORS[s]
        c1 = ws_sum.cell(row=row, column=1, value=STATUS_LABELS[s])
        c2 = ws_sum.cell(row=row, column=2, value=count)
        for c in (c1, c2):
            c.fill = PatternFill("solid", fgColor=bg_hex.lstrip("#"))
            c.font = Font(color=text_hex.lstrip("#"), bold=True)
        row += 1

    row += 1
    ws_sum.cell(row=row, column=1, value="By priority").font = Font(bold=True, size=11)
    row += 1
    ws_sum.cell(row=row, column=1, value="Priority").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1
    for p in PRIORITY_ORDER:
        count = by_priority.get(p, 0)
        if not count:
            continue
        ws_sum.cell(row=row, column=1, value=p.title())
        ws_sum.cell(row=row, column=2, value=count)
        row += 1

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 12

    # ---------- Tickets sheet ----------
    ws = wb.create_sheet("Tickets")
    ws.append(HEADER)
    _style_header_row(ws, 1, len(HEADER))
    ws.freeze_panes = "A2"

    thin = Border(bottom=Side(style="thin", color=GRID.lstrip("#")))

    for t in tickets:
        raw = _row(t)
        status_key = raw[1]
        display_row = [raw[0], STATUS_LABELS.get(status_key, status_key)] + raw[2:]
        ws.append(display_row)
        r = ws.max_row
        text_hex, bg_hex = STATUS_COLORS.get(status_key, ("#334155", "#F1F5F9"))
        status_cell = ws.cell(row=r, column=2)
        status_cell.fill = PatternFill("solid", fgColor=bg_hex.lstrip("#"))
        status_cell.font = Font(color=text_hex.lstrip("#"), bold=True)
        title_cell = ws.cell(row=r, column=1)
        title_cell.alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, len(HEADER) + 1):
            ws.cell(row=r, column=c).border = thin

    if tickets:
        last_row = ws.max_row
        last_col_letter = get_column_letter(len(HEADER))
        table_range = f"A1:{last_col_letter}{last_row}"
        xl_table = XLTable(displayName="TicketsTable", ref=table_range)
        xl_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
        )
        ws.add_table(xl_table)
    else:
        ws.cell(row=2, column=1, value="No tickets match these filters.").font = Font(italic=True, color=MUTED.lstrip("#"))

    widths = {1: 42, 2: 16, 3: 10, 4: 20, 5: 16, 6: 16, 7: 20, 8: 20, 9: 17, 10: 17}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
